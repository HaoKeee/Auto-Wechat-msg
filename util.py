#coding:utf-8
import datetime
import pyperclip
import alchemy as db
from config import global_config
from openpyxl import load_workbook
from PyQt5 import QtCore,QtWidgets

def copy_file(url):
  """
  将文件添加到粘贴板
  url:文件路径
  """
  app = QtWidgets.QApplication([])
  data = QtCore.QMimeData()
  url = QtCore.QUrl.fromLocalFile(url)
  data.setUrls([url])
  app.clipboard().setMimeData(data)

def copy_text(text):
  """
  将文本添加到粘贴板
  text:文本
  """
  pyperclip.copy(text)

def get_all_groups(root):
  """
  获取全部要发送的微信群列表
  root:所要发送的excel文件部分路径
  """
  today = datetime.datetime.today().strftime('%Y%m%d')
  groups = db.db_session.query(db.Groups).filter(db.Groups.send == True).all()
  return [{'location': group.location, 'name': group.name, 'path': root + group.path + today + '.xlsx'} for group in groups]

def read_xlsx_info(path):
  """
  获取excel文件的部分信息并返回统计值
  path:excel文件路径
  """
  wb = load_workbook(path)
  ws = wb['住宅']
  total_num = ws.max_row
  discount_num = 0
  for row in ws:
    discount = row[4].value
    try:
      if isinstance(discount, str):
        discount = discount.replace('%','')
        discount = float(discount)
        if discount <= 80:
          discount_num += 1
    except:
      continue
  return {
    'total': total_num,
    'discount': discount_num
  }

def get_text(location,counts):
  """
  根据地名和excel统计值拼接文本
  location:地名
  counts:excel统计值对象
  """
  return f'@所有人 大家好！\r今日{location}新推出{counts["total"]}套房地产项目，其中8折以下的项目{counts["discount"]}套，详见文件及图片，另提供房源原始Excel文件，可查看股权/债权和其他房产来源。\r打折资产，一网无遗。欢迎推荐大型机构专家入群。https://www.zichanxinxi.com  微信小程序:资产信息'
